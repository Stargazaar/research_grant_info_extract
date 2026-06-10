"""Load research-grant rows from the Excel workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

DEFAULT_PATH = Path("data/research_grants.xlsx")

# Map of spreadsheet headers -> GrantState field names.
COLUMN_MAP = {
    "Project ID": "project_id",
    "Project Title": "project_title",
    "Principal Investigator": "principal_investigator",
    "Host Institution": "host_institution",
    "Abstract": "abstract",
}


def _clean(value: object) -> str:
    """Normalise a cell value to a stripped string."""
    if value is None:
        return ""
    return str(value).strip()


def load_grants(path: str | Path = DEFAULT_PATH) -> list[dict[str, str]]:
    """Read the workbook and return one dict per grant row."""
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [_clean(h) for h in rows[0]]
    grants: list[dict[str, str]] = []
    for raw in rows[1:]:
        record = {header[i]: _clean(v) for i, v in enumerate(raw)}
        if not record.get("Abstract"):
            continue
        grants.append({field: record.get(col, "") for col, field in COLUMN_MAP.items()})

    wb.close()
    return grants
